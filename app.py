"""
title: 'atualiza-planilha-com-dados-web'
author: 'Elias Albuquerque'
version: 'Python 3.12.0'
created: '2024-05-20'
update: '2024-06-20'
"""


import openpyxl
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

# abrir planilha e coletar cpf
wb = openpyxl.load_workbook(filename='python-status_pagamento.xlsx')
ws = wb['Sheet1']
cpfs = []

for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
    cell = row[0].value  # Acessando a célula diretamente
    if cell is not None:  # Verificando se a célula tem valor
        # print(type(cell.value))
        if not isinstance(cell, str):
            int_data = int(cell)
            cell = str(int_data)
            
        data = re.sub(r'\D', '', cell)
        cpfs.append(data)

        # Reescrevendo o valor na célula
        row[0].value = data  # Coluna C (índice 0)

# Salva os dados reescritos
wb.save('python-status_pagamento.xlsx')  

# acessa o site
driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')
sleep(8)

input_cpf = '//input[@id="cpfInput"]'
consult = '//button[@type="submit"]'
status_label = '//span[@id="statusLabel"]'
payment_label = '//p[@id="paymentDate"]' 
method_label = '//p[@id="paymentMethod"]' 

results = {}

for cpf in cpfs:
    # insere o cpf e pesquisa
    digite_cpf = driver.find_element(By.XPATH, input_cpf)
    sleep(1)
    digite_cpf.clear()
    sleep(1)
    digite_cpf.send_keys(cpf)
    sleep(1)

    pesquisar = driver.find_element(By.XPATH, consult)
    sleep(1)
    pesquisar.click()
    sleep(5)

    # coleta os dados da resposta
    status_element = driver.find_element(By.XPATH, status_label)
    status = status_element.text
    sleep(1)

    if status == 'em dia':
        payment_date = driver.find_element(By.XPATH, payment_label)
        date = payment_date.text.split(': ')[1]
        sleep(.5)

        payment_method = driver.find_element(By.XPATH, method_label)
        method = payment_method.text.split(': ')[1]
        sleep(.5)
        
        results[cpf] = {
            'status': status,
            'data_pagamento': date,
            'metodo_pagamento': method,
        }
    elif status == 'atrasado':
        results[cpf] = {
            'status': status,
            'data_pagamento': '-',
            'metodo_pagamento': '-',
        }

# Verificar se a coluna E possui o cabeçalho "Status"
if ws['E1'].value != 'Status':
    ws['E1'] = 'Status'
    ws['F1'] = 'Data de Pagamento'
    ws['G1'] = 'Método de Pagamento'

# Inserir os dados do dicionário para cada CPF
for row in ws.iter_rows(min_row=2, max_col=7):  # Iterar até a coluna G (índice 6)
    cpf = row[2].value  # Coluna C (índice 2)
    if cpf in results:
        result = results[cpf]
        row[4].value = result['status']  # Coluna E (índice 4)
        row[5].value = result['data_pagamento']  # Coluna F (índice 5)
        row[6].value = result['metodo_pagamento']  # Coluna G (índice 6)

# Salvar as alterações na planilha
wb.save('python-status_pagamento.xlsx')